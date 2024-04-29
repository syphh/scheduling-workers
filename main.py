import openpyxl
import datetime as dt
import plotly.graph_objects as go
import plotly.express as px
import solver

HOURS_IN_DAY = 24
DAYS_IN_WEEK = 7
start_date = dt.datetime(2024, 4, 29)  # has to be a Monday


def plot_staffed_vs_required(staffed: list[int], requirements: list[list[int]]):
    weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    requirements_flattened = [item for row in requirements for item in row]
    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=list(range(len(staffed))), 
            y=staffed, 
            mode='lines', 
            name='Staffed'
        )
    )
    fig.add_trace(
        go.Scatter(
            x=list(range(len(requirements_flattened))), 
            y=requirements_flattened, 
            mode='lines', 
            name='Required'
        )
    )
    TICKS_HOURS_STEP = 6
    TICKS_PER_DAY = HOURS_IN_DAY // TICKS_HOURS_STEP
    fig.update_xaxes(
        tickmode='array',
        tickvals=list(range(0, DAYS_IN_WEEK*HOURS_IN_DAY+1, TICKS_HOURS_STEP)),
        ticktext=[
            f"{weekdays[(TICKS_HOURS_STEP*i)//HOURS_IN_DAY%DAYS_IN_WEEK]} {(TICKS_HOURS_STEP*i)%HOURS_IN_DAY}:00" 
            for i in range(DAYS_IN_WEEK*TICKS_PER_DAY+1)
        ],
        tickangle=80
    )
    fig.update_layout(
        title='Staffed vs Required', 
        xaxis_title='Hour of the week', 
        yaxis_title='Number of workers'
    )
    return fig


def plot_timeline(selected_shifts: list[tuple[dt.time, dt.time, list[str]]]):
    weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    data = []
    for i, (start_time, end_time, days_off) in enumerate(selected_shifts):
        for d in range(len(weekdays)):
            if weekdays[d] not in days_off:
                next_day_flag = int(end_time < start_time)
                start = start_date + dt.timedelta(
                    days=d, hours=start_time.hour, minutes=start_time.minute
                )
                end = start_date + dt.timedelta(
                    days=d + next_day_flag, hours=end_time.hour, minutes=end_time.minute
                )
                data.append(dict(Worker=f'Worker {i+1}', Start=start, Finish=end))
                if start.weekday() == 6 and end.weekday() == 0 and end.hour != 0:
                    data.append(
                        dict(
                            Worker=f'Worker {i+1}', 
                            Start=start - dt.timedelta(days=DAYS_IN_WEEK), 
                            Finish=end - dt.timedelta(days=DAYS_IN_WEEK)
                        )
                    )
    fig = px.timeline(data, x_start='Start', x_end='Finish', y='Worker')
    for i in range(DAYS_IN_WEEK + 1):
        fig.add_vline(x=start_date + dt.timedelta(days=i),  line_color='black')
    fig.update_yaxes(categoryorder='array', categoryarray=list(reversed(fig.data[0].y)))
    TICKS_HOURS_STEP = 6
    TICKS_PER_DAY = HOURS_IN_DAY // TICKS_HOURS_STEP
    fig.update_xaxes(
        tickmode='array',
        tickvals=[start_date + dt.timedelta(hours=TICKS_HOURS_STEP*i) for i in range(DAYS_IN_WEEK*TICKS_PER_DAY+1)],
        ticktext=[
            f"{weekdays[(TICKS_HOURS_STEP*i)//HOURS_IN_DAY%DAYS_IN_WEEK]} {(TICKS_HOURS_STEP*i)%HOURS_IN_DAY}:00" 
            for i in range(DAYS_IN_WEEK*TICKS_PER_DAY+1)
        ],
        tickangle=80
    )
    return fig


def read_input(filepath):
    wb = openpyxl.load_workbook(filepath)

    shifts_sheet = wb['Shifts']
    shift_types = []
    for row in shifts_sheet.iter_rows(min_row=2, values_only=True):
        assert isinstance(row[0], dt.time), \
            f"Shift start time should be a time object. Current: {row[0]}"
        assert isinstance(row[1], dt.time), \
            f"Shift end time should be a time object. Current: {row[1]}"
        assert isinstance(row[2], int), \
            f"Shift num days off should be an integer. Current: {row[2]}"
        assert 0 <= row[2] <= DAYS_IN_WEEK, \
            f"Shift num days off should be between 0 and 7. Current: {row[2]}"
        shift_types.append(row)
    
    requirements_sheet = wb['Requirements']
    requirements = []
    for row in requirements_sheet.iter_rows(min_row=2, values_only=True):
        assert all(isinstance(cell, int) for cell in row[1:]), \
            f"Requirement values should be integers."
        assert all(cell >= 0 for cell in row[1:]), \
            f"Requirement values should be non-negative."
        requirements.append(row[1:])
    
    return shift_types, requirements


if __name__ == '__main__':

    filepath_input = 'input.xlsx'
    filepath_output = 'output.xlsx'

    try:
        shift_types, requirements = read_input(filepath_input)
    except Exception as e:
        print(f"Error reading input: {e}")
        exit()
    
    try:
        staffed, selected_shifts = solver.recommend_shifts(shift_types, requirements)
        if not selected_shifts:
            print("No solution found.")
            exit()
    except Exception as e:
        print(f"Error solving the problem: {e}")
        exit()
    
    try:
        fig_staffed_vs_required = plot_staffed_vs_required(staffed, requirements)
        fig_timeline = plot_timeline(selected_shifts)
        fig_staffed_vs_required.write_html('staffed_vs_required.html', auto_open=True)
        fig_timeline.write_html('timeline.html', auto_open=True)
    except Exception as e:
        print(f"Error plotting the results: {e}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Workers'
    ws.append(['Worker ID', 'Start time', 'End time', 'Days off'])
    for i, (start_time, end_time, days_off) in enumerate(selected_shifts):
        ws.append([f'Worker {i+1}', start_time.strftime('%H:%M'), end_time.strftime('%H:%M'), ', '.join(days_off)])
    for i in range(1, len(selected_shifts)+2):
        for j in range(1, 5):
            ws.cell(row=i, column=j).font = openpyxl.styles.Font(bold=i == 1)
            ws.cell(row=i, column=j).fill = openpyxl.styles.PatternFill(
                start_color='A2E1E8', end_color='A2E1E8', fill_type='solid'
            )
            ws.cell(row=i, column=j).border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'), 
                right=openpyxl.styles.Side(style='thin'), 
                top=openpyxl.styles.Side(style='thin'), 
                bottom=openpyxl.styles.Side(style='thin')
            )
            ws.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center')
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = [10, 10, 10, 40][j-1]

    ws = wb.create_sheet('Staffed')
    ws.append(['Weekday', 'Hour', 'Staffed', 'Required'])
    weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    requirements_flattened = [item for row in requirements for item in row]
    for i in range(len(staffed)):
        ws.append([weekdays[i//HOURS_IN_DAY], dt.time(i%HOURS_IN_DAY).strftime('%H:%M'), staffed[i], requirements_flattened[i]])
    for i in range(1, len(staffed)+2):
        for j in range(1, 5):
            ws.cell(row=i, column=j).font = openpyxl.styles.Font(bold=i == 1)
            ws.cell(row=i, column=j).fill = openpyxl.styles.PatternFill(
                start_color='F2BDEF', end_color='F2BDEF', fill_type='solid'
            )
            ws.cell(row=i, column=j).border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'), 
                right=openpyxl.styles.Side(style='thin'), 
                top=openpyxl.styles.Side(style='thin'), 
                bottom=openpyxl.styles.Side(style='thin')
            )
            ws.cell(row=i, column=j).alignment = openpyxl.styles.Alignment(horizontal='center')
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 10
    
    try:
        wb.save(filepath_output)
    except Exception as e:
        print(f"Error saving the output: {e}")
        print("Please close the output file if it is open.")
        exit()
